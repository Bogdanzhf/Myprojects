import openpyxl
from openpyxl import Workbook
import os
import json
from datetime import datetime, timedelta
import random
from random import randint

# ------------------------ Угадай число/слово ------------------------
def guess_number():
    print("Угадываем число")
    print("Кто будет отгадывать? ВЫ - 1, Я - 2 ")
    b = int(input())
    if b == 1:
        Number = randint(1, 100)
        print("Угадайте целое число от 1 до 100:")
        Guess = int(input())
        while Guess != Number:
            print("Повторите попытку:")
            Guess = int(input())
            if Guess < Number:
                print("Ваше число меньше, чем задумал компьютер")
            elif Guess > Number:
                print("Ваше число больше, чем задумал компьютер")
            else:
                print("Вы угадали")
    else:
        print("Загадайте число от 1 до 100")
        begin = 1
        end = 100
        turn = 1
        while True:
            middle = (begin + end) // 2
            print(middle)
            utochnim = input(">, < или = ? ")
            if utochnim == '>':
                begin = middle + 1
            if utochnim == '<':
                end = middle - 1
            if utochnim == '=':
                print("Ура, я угадал за", turn, "попыток! :)")
                break
            turn += 1

def guess_word():
    words = ['компьютер', 'питон', 'программа', 'учеба', 'код', 'осень', 'привет', 'слово', 'вода', 'кот', 'собака', 'стол']
    word = random.choice(words)
    print("Угадайте буквы")
    guesses = ''
    turns = 12
    while turns > 0:
        failed = 0
        for bukva in word:
            if bukva in guesses:
                print(bukva, end=" ")
            else:
                print("_", end=" ")
                failed += 1
        if failed == 0:
            print("\nВы выиграли!")
            print("Слово: ", word)
            break
        print()
        guess = input("Угадайте букву: ")
        guesses += guess
        if guess not in word:
            turns -= 1
            print("Неверно")
            print(f"У вас есть ещё {turns} попыток")
            print()
        if turns == 0:
            print("Вы проиграли")

# ------------------------ Адресная книга ------------------------
class AddressBook:
    def __init__(self, filename='contacts.xlsx'):
        self.filename = filename
        self.contacts = self.load_contacts()

    def load_contacts(self):
        if os.path.exists(self.filename):
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb.active
            contacts = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                phone_number, name, blocked, favorite = row
                contacts[phone_number] = {
                    'name': name,
                    'blocked': blocked,
                    'favorite': favorite
                }
            return contacts
        else:
            return {}

    def save_contacts(self):
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Номер телефона", "Имя", "Заблокирован", "Избранное"])

        for phone_number, info in self.contacts.items():
            sheet.append([phone_number, info['name'], info['blocked'], info['favorite']])

        wb.save(self.filename)

    def create_contact(self):
        phone_number = input("Введите номер телефона: ")
        if phone_number in self.contacts:
            print(f"Контакт с номером {phone_number} уже существует.")
        else:
            name = input("Введите имя контакта: ")
            self.contacts[phone_number] = {'name': name, 'blocked': False, 'favorite': False}
            self.save_contacts()
            print(f"Контакт '{name}' с номером {phone_number} создан.")

    def call_contact(self):
        phone_number = input("Введите номер телефона контакта, которому хотите позвонить: ")
        if phone_number in self.contacts:
            name = self.contacts[phone_number]['name']
            print(f"Производится звонок контакту {name} ({phone_number})...")
        else:
            print(f"Контакт с номером {phone_number} не найден.")

    def delete_contact(self):
        phone_number = input("Введите номер телефона контакта, которого хотите удалить: ")
        if phone_number in self.contacts:
            name = self.contacts[phone_number]['name']
            confirm = input(f"Вы хотите удалить контакт {name} ({phone_number})? (да/нет): ").strip().lower()
            if confirm == 'да':
                del self.contacts[phone_number]
                self.save_contacts()
                print(f"Контакт '{name}' удален.")
            else:
                print("Удаление отменено.")
        else:
            print(f"Контакт с номером {phone_number} не найден.")

    def block_contact(self):
        phone_number = input("Введите номер телефона контакта, которого хотите заблокировать: ")
        if phone_number in self.contacts:
            name = self.contacts[phone_number]['name']
            confirm = input(f"Вы хотите заблокировать контакт {name} ({phone_number})? (да/нет): ").strip().lower()
            if confirm == 'да':
                self.contacts[phone_number]['blocked'] = True
                self.save_contacts()
                print(f"Контакт '{name}' заблокирован.")
            else:
                print("Блокировка отменена.")
        else:
            print(f"Контакт с номером {phone_number} не найден.")

    def write_to_contact(self):
        phone_number = input("Введите номер телефона контакта, которому хотите написать: ")
        if phone_number in self.contacts:
            name = self.contacts[phone_number]['name']
            message = input("Что вы хотите написать? ")
            print(f"Сообщение для {name} ({phone_number}): {message}")
        else:
            print(f"Контакт с номером {phone_number} не найден.")

    def add_favorite(self):
        phone_number = input("Введите номер телефона контакта, которого хотите добавить в избранное: ")
        if phone_number in self.contacts:
            name = self.contacts[phone_number]['name']
            self.contacts[phone_number]['favorite'] = True
            self.save_contacts()
            print(f"Контакт '{name}' добавлен в избранное.")
        else:
            print(f"Контакт с номером {phone_number} не найден.")

    def rename_contact(self):
        phone_number = input("Введите номер телефона контакта, которого хотите переименовать: ")
        if phone_number in self.contacts:
            old_name = self.contacts[phone_number]['name']
            new_name = input(f"Введите новое имя для контакта '{old_name}': ")
            self.contacts[phone_number]['name'] = new_name
            self.save_contacts()
            print(f"Контакт '{old_name}' переименован в '{new_name}'.")
        else:
            print(f"Контакт с номером {phone_number} не найден.")

    def display_contacts(self):
        if not self.contacts:
            print("Адресная книга пуста.")
        else:
            for phone_number, info in self.contacts.items():
                name = info['name']
                status = "заблокирован" if info['blocked'] else "активен"
                favorite = "избранное" if info['favorite'] else "обычный"
                print(f"Контакт: {name} ({phone_number}), Статус: {status}, Тип: {favorite}")

    def run(self):
        while True:
            print("\nВыберите действие:")
            print("1. Создать контакт")
            print("2. Позвонить контакту")
            print("3. Удалить контакт")
            print("4. Заблокировать контакт")
            print("5. Написать контакту")
            print("6. Добавить контакт в избранное")
            print("7. Переименовать контакт")
            print("8. Показать все контакты")
            print("9. Выйти в главное меню")

            choice = input("Введите номер действия: ").strip()
            if choice == '1':
                self.create_contact()
            elif choice == '2':
                self.call_contact()
            elif choice == '3':
                self.delete_contact()
            elif choice == '4':
                self.block_contact()
            elif choice == '5':
                self.write_to_contact()
            elif choice == '6':
                self.add_favorite()
            elif choice == '7':
                self.rename_contact()
            elif choice == '8':
                self.display_contacts()
            elif choice == '9':
                break
            else:
                print("Неверный выбор, попробуйте снова.")


# ------------------------ Расходы ------------------------
def load_expenses():
    """Загружает расходы из файла."""
    try:
        with open('expenses.json', 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return []  # Если файл не найден, возвращаем пустой список
    except json.JSONDecodeError:
        return []  # Если файл пуст или поврежден, возвращаем пустой список

def save_expenses(expenses):
    """Сохраняет расходы в файл."""
    with open('expenses.json', 'w') as file:
        json.dump(expenses, file)

def add_expense(expenses):
    """Добавляет новый расход."""
    name = input("Введите название расхода: ")
    amount = float(input("Введите сумму расхода: "))
    date = input("Введите дату (в формате дд.мм.гггг, оставьте пустым для сегодняшней даты): ")
    if not date:  # Если дата не введена, используем сегодняшнюю
        date = datetime.now().strftime('%d.%m.%Y')
    else:
        # Проверяем правильность введенной даты
        try:
            datetime.strptime(date, '%d.%m.%Y')  # Проверка формата даты
        except ValueError:
            print("Неверный формат даты. Используйте дд.мм.гггг.")
            return
    
    expenses.append({
        'name': name,
        'amount': amount,
        'date': date
    })
    print("Расход добавлен.")

def view_expenses(expenses):
    """Просматривает все расходы."""
    if not expenses:
        print("Нет расходов для отображения.")
        return
    
    print("\nСписок расходов:")
    for index, expense in enumerate(expenses, start=1):  # Нумерация с 1
        print(f"{index}. {expense['date']} - {expense['name']}: {expense['amount']}")

def delete_expense(expenses):
    """Удаляет расход по индексу."""
    view_expenses(expenses)
    if not expenses:
        return
    
    try:
        index = int(input("Введите номер расхода для удаления: ")) - 1  # Приводим к индексу массива
        if 0 <= index < len(expenses):
            removed = expenses.pop(index)
            print(f"Расход '{removed['name']}' удалён.")
        else:
            print("Неверный индекс.")
    except ValueError:
        print("Пожалуйста, введите корректный номер.")

def calculate_total(expenses):
    """Подсчитывает общую сумму расходов."""
    return sum(expense['amount'] for expense in expenses)

def run_expenses():
    expenses = load_expenses()
    
    while True:
        print("\nМеню:")
        print("1. Добавить расход")
        print("2. Просмотреть расходы")
        print("3. Удалить расход")
        print("4. Посчитать общую сумму расходов")
        print("5. Выйти в главное меню")

        choice = input("Выберите действие (1-5): ")
        
        if choice == '1':
            add_expense(expenses)
            save_expenses(expenses)
        elif choice == '2':
            view_expenses(expenses)
        elif choice == '3':
            delete_expense(expenses)
            save_expenses(expenses)
        elif choice == '4':
            total = calculate_total(expenses)
            print(f"Общая сумма расходов: {total}")
        elif choice == '5':
            print("Выход из программы.")
            break
        else:
            print("Неверный ввод. Пожалуйста, попробуйте снова.")


# ------------------------ Юридический калькулятор ------------------------
def calculate_claim_deadline(start_date, claim_type):
    # Конвертируем строку в дату
    start_date = datetime.strptime(start_date, '%d-%m-%Y')

    # Определим сроки для различных типов гражданских исков
    claim_periods = {
        "по договорам": 1,  # срок по договорам - 1 год
        "по возмещению ущерба": 2,  # срок по возмещению ущерба - 2 года
        "по обязательствам": 10  # по долговым обязательствам - 10 лет
    }

    # Получим срок для данного типа иска
    period = claim_periods.get(claim_type.lower())

    if not period:
        raise ValueError(f"Неизвестный тип иска: {claim_type}")

    # Рассчитаем дату окончания срока
    deadline_date = start_date + timedelta(days=period * 365)  # Приблизительно 365 дней в году

    # Проверим, если срок заканчивается в выходные или праздничные дни, перенесем на следующий рабочий день
    while deadline_date.weekday() >= 5:  # Если суббота (5) или воскресенье (6)
        deadline_date += timedelta(days=1)

    return deadline_date.strftime('%d-%m-%Y')

def show_claim_types():
    print("Доступные типы гражданских исков:")
    print("1. По договорам")
    print("2. По возмещению ущерба")
    print("3. По обязательствам")

def save_result_to_file(start_date, claim_type, deadline):
    result = {
        "start_date": start_date,
        "claim_type": claim_type,
        "deadline": deadline
    }
    
    with open('claim_results.json', 'a') as file:
        file.write(json.dumps(result) + '\n')  # Сохраняем результат в формате JSON

def run_legal_calculator():
    print("Добро пожаловать в юридический калькулятор исковых сроков!")
    show_claim_types()  # Показать доступные типы исков
    
    # Получаем данные от пользователя
    start_date = input("Введите дату начала срока (дд-мм-гггг): ")
    claim_type = input("Введите тип иска (по договорам, по возмещению ущерба, по обязательствам): ")
    
    try:
        deadline = calculate_claim_deadline(start_date, claim_type)
        print(f"Дата окончания срока подачи иска: {deadline}")
        
        # Сохраняем результат в файл
        save_result_to_file(start_date, claim_type, deadline)
        
    except ValueError as e:
        print(f"Ошибка: {e}")


# ------------------------ Рецепты ------------------------
ingredients_list = [
    "Картофель", "Макароны", "Морковь", "Лук", "Фарш", 
    "Сосиски", "Колбаски", "Ветчина", "Куриные бедрышки", "Куриное филе",
    "Курица", "Сыр", "Тыква", "Зелень", "Помидоры", 
    "Сметана", "Яйца", "Кабачки"
]

dishes = {
    "Ленивая лазанья с куриным фаршем, овощами и сыром": 
        ["фарш", "макароны", "лук", "морковь", "молоко", "томатная паста", "сыр", "соль", "паприка", "масло"],
    "Макароны в сливочном соусе": 
        ["макароны", "ветчина", "сыр", "сливки", "чеснок", "масло", "соль", "зелень"],
    "Макароны в сливочно-томатном соусе с колбасками": 
        ["макароны", "колбаски", "помидор", "сливки", "сыр", "лук", "чеснок", "соль", "перец", "базилик", "масло"],
    "Картошка по-деревенски в духовке": 
        ["картофель", "масло", "соль", "зелень", "пряности"],
    "Картофельные драники без яиц и муки": 
        ["картофель", "лук", "соль", "масло", "сметана"],
    "Пестрый суп": 
        ["картофель", "лук", "морковь", "масло", "консервированная кукуруза", "вермишель", "сосиски", "соль", "перец", "карри", "зелень"],
    "Кабачки с соевым соусом и чесноком на сковороде": 
        ["кабачки", "чеснок", "соевый соус", "масло", "крахмал", "укроп", "кунжут", "паприка", "сахар", "соль"],
    "Запеканка из кабачков с курицей и сыром": 
        ["кабачки", "куриное филе", "лук", "яйцо", "сыр", "сметана", "масло", "горчица", "соль"],
    "Кабачки запеченные с фаршем и рисом": 
        ["кабачки", "фарш", "рис", "помидоры", "сыр", "перец", "соль", "сметана"],
    "Пирог из тыквы": 
        ["тыква", "яйца", "сметана", "масло", "сахар", "мука", "соль", "ванильный сахар", "разрыхлитель"],
    "Тыква по-гречески": 
        ["тыква", "чеснок", "томатная паста", "базилик", "масло", "вода", "соль", "перец"],
    "Тыквенные панкейки": 
        ["тыква", "яйца", "масло", "молоко", "мука", "сахар", "разрыхлитель", "соль"],
    "Чахохбили из курицы": 
        ["куриные бедрышки", "помидоры", "лук", "масло", "хмели-сунели", "соль", "перец", "чеснок", "кинза", "базилик"],
    "Куриные рулетики с сыром": 
        ["куриное филе", "сыр", "зелень", "чеснок", "соль", "перец", "мука", "масло"],
    "Куриный суп с чесночными галушками": 
        ["курица", "картофель", "морковь", "лук", "лавровый лист", "вода", "соль", "перец", "мука", "яйцо", "масло"]
}

# Функция для сохранения данных в файл в формате JSON
def save_data_to_file(filename, data):
    # Открываем файл для записи
    with open(filename, 'w', encoding='utf-8') as file:
        # Сохраняем данные в формате JSON
        json.dump(data, file, ensure_ascii=False, indent=4)

# Функция для загрузки данных из файла в формате JSON
def load_data_from_file(filename):
    """Загрузка данных из файла в формате JSON."""
    with open(filename, 'r', encoding='utf-8') as file:
        return json.load(file)
    
# Функция для получения ингредиентов от пользователя
def get_user_ingredients():
    """Получение ингредиентов от пользователя."""
    print("Список ингредиентов:")
    for index, ingredient in enumerate(ingredients_list, start=1):
        print(f"{index}. {ingredient}")
    
    # Запрашиваем ингредиенты у пользователя и превращаем их в множество
    user_ingredients = input("Введите ингредиенты, которые у вас есть (через запятую): ").lower().split(',')
    return {ingredient.strip() for ingredient in user_ingredients}

# Функция для поиска доступных блюд
def find_available_dishes(user_ingredients):
    """Поиск блюд, которые могут быть приготовлены из имеющихся ингредиентов."""
    available_dishes = {}
    for dish, ingredients in dishes.items():
        if set(ingredients).intersection(user_ingredients):
            available_dishes[dish] = ingredients
    return available_dishes

# Функция для выбора блюда пользователем
def choose_dish(available_dishes):
    """Выбор блюда пользователем."""
    print("\nДоступные блюда:")
    for index, dish in enumerate(available_dishes.keys(), start=1):
        print(f"{index}. {dish}")
    
    choice = int(input("Выберите номер блюда, которое хотите приготовить: "))
    selected_dish = list(available_dishes.keys())[choice - 1]
    return selected_dish, available_dishes[selected_dish]

# Функция для запуска рецептов
def run_recipes():
    user_ingredients = get_user_ingredients()
    available_dishes = find_available_dishes(user_ingredients)

    if not available_dishes:
        print("К сожалению, ни одно блюдо не может быть приготовлено с вашими ингредиентами.")
        return
    
    selected_dish, ingredients = choose_dish(available_dishes)
    print(f"\nВы выбрали: {selected_dish}")
    print("Ингредиенты для этого блюда:")
    for ingredient in ingredients:
        print(f"- {ingredient.capitalize()}")

    # Сохранение данных в файл
    save_data_to_file('available_dishes.json', available_dishes)
    print(f"\nДанные о доступных блюдах сохранены в файле 'available_dishes.json'.")

# ------------------------ Главное меню ------------------------
def main_menu():
    while True:
        print("\nВыберите приложение:")
        print("1. Угадай число")
        print("2. Угадай слово")
        print("3. Адресная книга")
        print("4. Расходы")
        print("5. Юридический калькулятор")
        print("6. Рецепты")
        print("0. Выход")

        choice = input("Введите номер приложения: ")

        if choice == '1':
            guess_number()
            input("Нажмите Enter, чтобы вернуться в меню...")
        elif choice == '2':
            guess_word()
            input("Нажмите Enter, чтобы вернуться в меню...")
        elif choice == '3':
            address_book = AddressBook()
            address_book.run()
        elif choice == '4':
            run_expenses()
        elif choice == '5':
            run_legal_calculator()
        elif choice == '6':
            run_recipes()
        elif choice == '0':
            print("Выход из программы.")
            break
        else:
            print("Неверный ввод, попробуйте снова.")

# Запуск основного меню
if __name__ == "__main__":
    main_menu() 